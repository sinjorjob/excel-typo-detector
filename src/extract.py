"""
Excelファイルからセルデータを抽出するモジュール
"""

import os
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from loguru import logger


@dataclass
class CellData:
    """セルデータを格納するデータクラス"""
    file_name: str
    sheet_name: str
    cell_address: str
    row: int
    column: int
    text: str


class ExcelExtractor:
    """Excel ファイルからセルデータを抽出するクラス"""
    
    def __init__(self, config: Dict[str, Any]):
        """
        初期化
        
        Args:
            config: 設定辞書
        """
        self.config = config
        self.processing_config = config.get('processing', {})
        self.exclusions_config = config.get('exclusions', {})
        
        # 除外パターンのコンパイル
        self.skip_cell_patterns = [
            re.compile(pattern) 
            for pattern in self.exclusions_config.get('skip_cell_patterns', [])
        ]
    
    def extract_from_directory(self, input_dir: str) -> List[CellData]:
        """
        指定ディレクトリ内のすべてのExcelファイルからセルデータを抽出
        
        Args:
            input_dir: 入力ディレクトリパス
            
        Returns:
            抽出されたセルデータのリスト
        """
        input_path = Path(input_dir)
        if not input_path.exists():
            logger.error(f"Input directory does not exist: {input_dir}")
            return []
            
        all_cells = []
        excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
        
        if not excel_files:
            logger.warning(f"No Excel files found in {input_dir}")
            return []
            
        logger.info(f"Found {len(excel_files)} Excel files")
        
        for file_path in excel_files:
            try:
                cells = self.extract_from_file(str(file_path))
                all_cells.extend(cells)
                logger.info(f"Extracted {len(cells)} cells from {file_path.name}")
            except Exception as e:
                logger.error(f"Error processing {file_path.name}: {e}")
                
        logger.info(f"Total extracted cells: {len(all_cells)}")
        return all_cells
    
    def extract_from_file(self, file_path: str) -> List[CellData]:
        """
        単一のExcelファイルからセルデータを抽出
        
        Args:
            file_path: Excelファイルパス
            
        Returns:
            抽出されたセルデータのリスト
        """
        file_name = Path(file_path).name
        cells = []
        
        try:
            # openpyxlを使用してワークブックを読み込み
            workbook = load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                # スキップするシート名のチェック
                if self._should_skip_sheet(sheet_name):
                    logger.debug(f"Skipping sheet: {sheet_name}")
                    continue
                    
                sheet_cells = self._extract_from_sheet(
                    workbook[sheet_name], file_name, sheet_name
                )
                cells.extend(sheet_cells)
                
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            
        return cells
    
    def _extract_from_sheet(self, sheet, file_name: str, sheet_name: str) -> List[CellData]:
        """
        単一のワークシートからセルデータを抽出
        
        Args:
            sheet: openpyxlのWorksheetオブジェクト
            file_name: ファイル名
            sheet_name: シート名
            
        Returns:
            抽出されたセルデータのリスト
        """
        cells = []
        max_cells = self.processing_config.get('max_cells_per_sheet', 10000)
        skip_empty = self.processing_config.get('skip_empty_cells', True)
        include_formulas = self.processing_config.get('include_formulas', False)
        
        cell_count = 0
        
        for row in sheet.iter_rows():
            if cell_count >= max_cells:
                logger.warning(f"Reached max cells limit ({max_cells}) for sheet {sheet_name}")
                break
                
            for cell in row:
                if cell_count >= max_cells:
                    break
                    
                # セルの値を取得
                cell_value = cell.value
                
                # 空セルのスキップ
                if skip_empty and (cell_value is None or str(cell_value).strip() == ""):
                    continue
                    
                # 数式セルのスキップ（オプション）
                if not include_formulas and str(cell_value).startswith('='):
                    continue
                    
                # 文字列に変換
                text = str(cell_value) if cell_value is not None else ""
                
                # 除外パターンのチェック
                if self._should_skip_cell_content(text):
                    continue
                
                # CellDataオブジェクトを作成
                cell_data = CellData(
                    file_name=file_name,
                    sheet_name=sheet_name,
                    cell_address=cell.coordinate,
                    row=cell.row,
                    column=cell.column,
                    text=text
                )
                
                cells.append(cell_data)
                cell_count += 1
        
        logger.debug(f"Sheet '{sheet_name}': extracted {len(cells)} cells")
        return cells
    
    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """
        シートをスキップするかどうかを判定
        
        Args:
            sheet_name: シート名
            
        Returns:
            スキップする場合True
        """
        skip_sheets = self.exclusions_config.get('skip_sheets', [])
        
        for pattern in skip_sheets:
            if pattern.lower() in sheet_name.lower():
                return True
                
        return False
    
    def _should_skip_cell_content(self, text: str) -> bool:
        """
        セル内容をスキップするかどうかを判定
        
        Args:
            text: セルの内容
            
        Returns:
            スキップする場合True
        """
        if not text or text.strip() == "":
            return True
            
        # 除外パターンとのマッチング
        for pattern in self.skip_cell_patterns:
            if pattern.match(text):
                return True
                
        return False
    
    def save_extracted_data(self, cells: List[CellData], output_path: str):
        """
        抽出したセルデータをCSVファイルに保存
        
        Args:
            cells: セルデータのリスト
            output_path: 出力ファイルパス
        """
        try:
            data = []
            for cell in cells:
                data.append({
                    'file_name': cell.file_name,
                    'sheet_name': cell.sheet_name,
                    'cell_address': cell.cell_address,
                    'row': cell.row,
                    'column': cell.column,
                    'text': cell.text
                })
            
            df = pd.DataFrame(data)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            logger.info(f"Extracted data saved to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving extracted data: {e}")


def load_sample_files(sample_dir: str = "sample_data") -> List[CellData]:
    """
    サンプルファイルを読み込む便利関数
    
    Args:
        sample_dir: サンプルディレクトリパス
        
    Returns:
        抽出されたセルデータのリスト
    """
    # デフォルト設定
    default_config = {
        'processing': {
            'max_cells_per_sheet': 10000,
            'skip_empty_cells': True,
            'include_formulas': False
        },
        'exclusions': {
            'skip_sheets': ['設定', 'Config', 'メモ', 'Notes'],
            'skip_cell_patterns': [
                r'^=.*',  # 数式セル
                r'^[0-9]+$',  # 数値のみ
                r'^[a-zA-Z0-9_.-]+@[a-zA-Z0-9_.-]+$'  # メールアドレス
            ]
        }
    }
    
    extractor = ExcelExtractor(default_config)
    return extractor.extract_from_directory(sample_dir)


if __name__ == "__main__":
    # テスト実行
    import yaml
    
    # 設定ファイルを読み込み
    with open("config.yml", "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    
    # 抽出実行
    extractor = ExcelExtractor(config)
    cells = extractor.extract_from_directory("sample_data")
    
    # 結果を表示
    print(f"Extracted {len(cells)} cells")
    for i, cell in enumerate(cells[:5]):  # 最初の5件を表示
        print(f"{i+1}: {cell.file_name} - {cell.sheet_name}:{cell.cell_address} = '{cell.text}'")