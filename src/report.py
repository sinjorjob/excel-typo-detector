"""
レポート出力モジュール
検出結果をExcelレポートとして出力する
"""

import os
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from loguru import logger

from .extract import CellData
# アルゴリズム検出関連は削除済み - LLMレビューのみ使用
from .llm import LLMReviewResponse


class ReportGenerator:
    """レポート生成クラス"""
    
    def __init__(self, config: Dict[str, Any], output_dir: str = "data/output"):
        """
        初期化
        
        Args:
            config: 設定辞書
            output_dir: 出力ディレクトリ
        """
        self.config = config
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # レポート設定
        self.report_config = config.get('report', {})
        
        # Accenture風スタイル定義
        self.styles = self._define_styles()
        
        # タイムスタンプ
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def _define_styles(self) -> Dict[str, Any]:
        """スタイル定義"""
        return {
            # Accenture Purple
            'header_fill': PatternFill(start_color="A100FF", end_color="A100FF", fill_type="solid"),
            'header_font': Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF"),
            'normal_font': Font(name="Meiryo UI", size=10),
            'title_font': Font(name="Meiryo UI", size=14, bold=True),
            
            # ボーダー
            'thin_border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            
            # 問題レベル別の色分け
            'variant_fill': PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"),
            'typo_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'high_confidence_fill': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),
            'low_confidence_fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            
            # 中央揃え
            'center_alignment': Alignment(horizontal='center', vertical='center'),
            'left_alignment': Alignment(horizontal='left', vertical='center'),
        }
    
    def generate_all_reports(
        self, 
        detection_results: List = None,  # 互換性のため残すが使用しない
        llm_results: Optional[List[LLMReviewResponse]] = None,
        statistics: Optional[Dict[str, Any]] = None
    ) -> Dict[str, str]:
        """
        すべてのレポートを生成（LLMレビューのみ）
        
        Args:
            detection_results: 互換性のため残すが使用しない
            llm_results: LLMレビュー結果のリスト
            statistics: 統計情報
            
        Returns:
            生成されたファイルパスの辞書
        """
        generated_files = {}
        
        try:
            # メインレビューワークブック（LLMレビューのみ）
            main_report_path = self.generate_main_report(
                [], llm_results, statistics
            )
            generated_files['main_report'] = main_report_path
            
            # 自動修正案ワークブック（無効化 - レビュー台帳に修正案が含まれるため不要）
            # auto_fix_path = self.generate_auto_fix_report(detection_results)
            # generated_files['auto_fix'] = auto_fix_path
            
            # LLMレビュー結果JSON
            if llm_results:
                llm_json_path = self.save_llm_results_json(llm_results)
                generated_files['llm_json'] = llm_json_path
            
            # 統計レポートJSON
            if statistics:
                stats_json_path = self.save_statistics_json(statistics)
                generated_files['statistics'] = stats_json_path
            
            logger.info(f"Generated {len(generated_files)} report files")
            return generated_files
            
        except Exception as e:
            logger.error(f"Error generating reports: {e}")
            return {}
    
    def generate_main_report(
        self, 
        detection_results: List = None,  # 互換性のため残すが使用しない
        llm_results: Optional[List[LLMReviewResponse]] = None,
        statistics: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        メインレビューワークブック（LLMレビューのみ）を生成
        
        Args:
            detection_results: 互換性のため残すが使用しない
            llm_results: LLMレビュー結果のリスト
            statistics: 統計情報
            
        Returns:
            生成されたファイルパス
        """
        wb = Workbook()
        
        # デフォルトシートを削除
        wb.remove(wb.active)
        
        # 1. サマリーシート（LLMレビューのみ）
        self._create_summary_sheet(wb, [], statistics)
        
        # 2. LLMレビューシート（メイン結果）
        if llm_results:
            self._create_llm_review_sheet(wb, llm_results, [])
        else:
            # LLMレビュー結果がない場合でも空シートを作成
            self._create_empty_llm_sheet(wb)
        
        # ファイル保存
        output_path = self.output_dir / f"レビュー台帳_{self.timestamp}.xlsx"
        wb.save(output_path)
        
        logger.info(f"Main report saved to {output_path}")
        return str(output_path)
    
    def _create_summary_sheet(
        self, 
        wb: Workbook, 
        results: List = None,  # 互換性のため残すが使用しない
        statistics: Optional[Dict[str, Any]] = None
    ):
        """サマリーシート（LLMレビューのみ）を作成"""
        ws = wb.create_sheet("サマリー", 0)
        
        # タイトル
        ws['A1'] = "LLM誤字脱字検出レポート"
        ws['A1'].font = self.styles['title_font']
        ws.merge_cells('A1:E1')
        
        # 基本情報
        row = 3
        ws[f'A{row}'] = "実行日時："
        ws[f'B{row}'] = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
        row += 1
        
        # 機械的検出とLLMレビューの統計情報
        if statistics:
            mechanical = statistics.get('mechanical_detection', {})
            llm_review = statistics.get('llm_review', {})
            llm_issue_counts = llm_review.get('issue_type_counts', {})
            mechanical_types = mechanical.get('detection_types', {})
            
            # 総検出数を計算
            total_detections = mechanical.get('total_detections', 0) + llm_review.get('total_reviews', 0)
            ws[f'A{row}'] = "総検出件数："
            ws[f'B{row}'] = total_detections
            row += 2
            
            # 統計情報テーブル
            stats_data = [
                ["項目", "件数"],
                ["機械的検出", ""],
                ["- 表記ゆれ", mechanical.get('total_detections', 0)],
                ["- 平均信頼度", f"{mechanical.get('avg_confidence', 0):.2f}"],
                ["LLM検出", ""],
                ["- 誤字", llm_issue_counts.get('typo', 0)],
                ["- 表記ゆれ", llm_issue_counts.get('variant', 0)],
                ["- 修正不要", llm_issue_counts.get('none', 0)],
                ["- 高信頼度 (≥0.9)", llm_review.get('high_confidence', 0)],
                ["- 中信頼度 (0.7-0.9)", llm_review.get('medium_confidence', 0)],
                ["- 低信頼度 (<0.7)", llm_review.get('low_confidence', 0)],
                ["- LLMリクエスト回数", llm_review.get('request_count', 0)]
            ]
            
            self._create_table(ws, stats_data, start_row=row, table_name="Stats")
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_empty_llm_sheet(self, wb: Workbook):
        """LLMレビュー結果がない場合の空シートを作成"""
        ws = wb.create_sheet("LLMレビュー結果")
        
        # ヘッダー作成
        headers = [
            "ファイル名", "シート名", "セル位置", "元テキスト", 
            "問題タイプ", "修正案", "理由", "信頼度"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center_alignment']
        
        # メッセージ
        ws['A2'] = "修正が必要な項目は見つかりませんでした"
        ws['A2'].font = Font(italic=True, color="666666")
        
        # 列幅調整
        for col, width in enumerate([20, 15, 10, 30, 12, 20, 30, 10], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 削除済み: _create_detection_results_sheet - LLMレビューに統一
    
    # 削除済み: _create_file_summary_sheet - LLMレビューに統一
    
    def _create_llm_review_sheet(self, wb: Workbook, mixed_results: List, detection_results: List = None):
        """混合結果（機械的検出 + LLM検出）のレビューシートを作成"""
        ws = wb.create_sheet("検出結果")
        
        headers = [
            "ファイル名", "シート名", "セル位置", "元テキスト", 
            "問題タイプ", "修正案", "正準表記", "理由", "信頼度", "検出方式"
        ]
        
        data = [headers]
        needs_fix_rows = []  # 修正が必要な行のインデックスを記録
        
        for idx, result in enumerate(mixed_results):
            # 機械的検出結果かLLM検出結果かを判定
            if hasattr(result, 'file_name'):
                # 機械的検出結果（DetectionResult型）
                row_data = [
                    result.file_name,
                    result.sheet_name,
                    result.cell_address,
                    result.original,
                    result.issue_type,
                    result.suggested_fix or "",
                    result.canonical or "",
                    result.reason,
                    f"{result.confidence:.2f}",
                    "機械的検出"
                ]
                
                # 機械的検出は基本的に修正が必要（修正案の有無に関係なく）
                if result.issue_type in ["typo", "variant"]:
                    needs_fix_rows.append(idx + 2)
                    
            else:
                # LLM検出結果（LLMReviewResponse型）
                if hasattr(result, 'source_detection') and result.source_detection:
                    detection = result.source_detection
                    row_data = [
                        detection.cell_data.file_name,
                        detection.cell_data.sheet_name,
                        detection.cell_data.cell_address,
                        result.original,
                        result.issue_type,
                        result.suggested_fix or "",
                        result.canonical or "",
                        result.reason,
                        f"{result.confidence:.2f}",
                        "LLM検出"
                    ]
                else:
                    # source_detectionがない場合の基本情報
                    row_data = [
                        "不明", "不明", "不明",
                        result.original,
                        result.issue_type,
                        result.suggested_fix or "",
                        result.canonical or "",
                        result.reason,
                        f"{result.confidence:.2f}",
                        "LLM検出"
                    ]
                
                # LLM検出結果の修正要否判定
                fix_suggestion = result.suggested_fix or result.canonical
                is_no_problem = "問題なし" in result.reason or "問題な し" in result.reason
                if (result.issue_type in ["typo", "variant"] and 
                    fix_suggestion and 
                    result.original != fix_suggestion and
                    not is_no_problem):
                    needs_fix_rows.append(idx + 2)  # ヘッダー行を考慮して +2
            
            data.append(row_data)
        
        self._create_table(ws, data, start_row=1, table_name="LLMReviews")
        
        # 修正が必要な行をハイライト表示
        self._apply_llm_review_formatting(ws, needs_fix_rows, len(data))
        
        # 列幅調整（検出方式列を追加）
        column_widths = [20, 15, 10, 25, 12, 25, 20, 40, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    # 削除済み: _create_issue_type_sheets, _create_issue_detail_sheet - LLMレビューに統一
    
    def generate_auto_fix_report(self, results: List = None) -> str:  # 削除予定
        """
        自動修正案ワークブックを生成
        
        Args:
            results: 検出結果のリスト
            
        Returns:
            生成されたファイルパス
        """
        # 自動修正可能な結果のみ抽出
        auto_fix_results = [r for r in results if r.auto_fix]
        
        if not auto_fix_results:
            logger.info("No auto-fixable issues found")
            return ""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "自動修正案"
        
        # ヘッダー作成
        headers = [
            "ファイル名", "シート名", "セル位置", "現在のテキスト", 
            "修正後テキスト", "修正理由", "信頼度"
        ]
        
        data = [headers]
        for result in auto_fix_results:
            row_data = [
                result.cell_data.file_name,
                result.cell_data.sheet_name,
                result.cell_data.cell_address,
                result.original,
                result.suggested_fix,
                result.reason,
                f"{result.confidence:.2f}"
            ]
            data.append(row_data)
        
        # テーブル作成
        self._create_table(ws, data, start_row=1, table_name="AutoFixes")
        
        # 列幅調整
        column_widths = [20, 15, 10, 30, 30, 40, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # ファイル保存
        output_path = self.output_dir / f"自動修正案_{self.timestamp}.xlsx"
        wb.save(output_path)
        
        logger.info(f"Auto-fix report saved to {output_path} ({len(auto_fix_results)} items)")
        return str(output_path)
    
    def _create_table(
        self, 
        ws, 
        data: List[List], 
        start_row: int = 1, 
        table_name: str = "Table"
    ):
        """Excelテーブルを作成"""
        if not data:
            return
        
        # データを書き込み
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # ヘッダー行のスタイル
                if row_idx == start_row:
                    cell.fill = self.styles['header_fill']
                    cell.font = self.styles['header_font']
                    cell.alignment = self.styles['center_alignment']
                else:
                    cell.font = self.styles['normal_font']
                    cell.alignment = self.styles['left_alignment']
                
                cell.border = self.styles['thin_border']
        
        # テーブル範囲
        end_row = start_row + len(data) - 1
        end_col = len(data[0])
        table_range = f"A{start_row}:{chr(64 + end_col)}{end_row}"
        
        # テーブル作成
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    def _apply_conditional_formatting(self, ws, results: List = None):  # 削除予定
        """条件付き書式を適用"""
        if not results:
            return
        
        # データ範囲
        max_row = len(results) + 1
        
        # 問題タイプ別の色分け（D列：問題タイプ）
        for i, result in enumerate(results, 2):
            if result.issue_type == IssueType.VARIANT:
                ws[f'D{i}'].fill = self.styles['variant_fill']
            elif result.issue_type == IssueType.TYPO:
                ws[f'D{i}'].fill = self.styles['typo_fill']
            
            # 信頼度別の色分け（H列：信頼度）
            if result.confidence >= 0.9:
                ws[f'H{i}'].fill = self.styles['high_confidence_fill']
            elif result.confidence < 0.7:
                ws[f'H{i}'].fill = self.styles['low_confidence_fill']
    
    def _apply_llm_review_formatting(self, ws, needs_fix_rows: List[int], total_rows: int):
        """LLMレビューシートの条件付き書式を適用（修正が必要な行をハイライト）"""
        if not needs_fix_rows:
            return
        
        # 修正が必要な行を赤色でハイライト
        highlight_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        highlight_font = Font(name="Meiryo UI", size=10, bold=True)
        
        for row_num in needs_fix_rows:
            if row_num <= total_rows:
                # 行全体をハイライト（A列からJ列まで）
                for col in range(1, 11):  # A-J列
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = highlight_fill
                    if col in [4, 5, 6]:  # 元テキスト、修正案、正準表記列を太字に
                        cell.font = highlight_font
    
    def save_llm_results_json(self, llm_results: List[LLMReviewResponse]) -> str:
        """LLMレビュー結果をJSONで保存"""
        output_path = self.output_dir / f"llm_review_results_{self.timestamp}.json"
        
        data = {
            'timestamp': datetime.now().isoformat(),
            'total_reviews': len(llm_results),
            'results': [result.to_dict() for result in llm_results]
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"LLM results saved to {output_path}")
        return str(output_path)
    
    def save_statistics_json(self, statistics: Dict[str, Any]) -> str:
        """統計情報をJSONで保存"""
        output_path = self.output_dir / f"statistics_{self.timestamp}.json"
        
        data = {
            'timestamp': datetime.now().isoformat(),
            'statistics': statistics
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Statistics saved to {output_path}")
        return str(output_path)
    
    def generate_csv_export(self, results: List = None) -> str:  # 削除予定
        """CSV形式でエクスポート"""
        output_path = self.output_dir / f"detection_results_{self.timestamp}.csv"
        
        data = []
        for result in results:
            data.append({
                'ファイル名': result.cell_data.file_name,
                'シート名': result.cell_data.sheet_name,
                'セル位置': result.cell_data.cell_address,
                '問題タイプ': result.issue_type.value,
                '元テキスト': result.original,
                '修正案': result.suggested_fix or "",
                '信頼度': result.confidence,
                '理由': result.reason,
                '自動修正': result.auto_fix,
                '文脈': result.context
            })
        
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        logger.info(f"CSV export saved to {output_path}")
        return str(output_path)


# def generate_sample_report(sample_dir: str = "sample_data") -> Dict[str, str]:
#     """
#     サンプルデータでレポートを生成する便利関数
#     ※ アルゴリズム検出機能削除のためコメントアウト
#     
#     Args:
#         sample_dir: サンプルデータディレクトリ
#         
#     Returns:
#         生成されたファイルパスの辞書
#     """
#     # アルゴリズム検出機能削除のためコメントアウト
#     pass


if __name__ == "__main__":
    # テスト実行
    # アルゴリズム検出機能削除のためコメントアウト
    # generated_files = generate_sample_report()
    # 
    # print("=== Generated Reports ===")
    # for report_type, file_path in generated_files.items():
    #     print(f"{report_type}: {file_path}")
    pass